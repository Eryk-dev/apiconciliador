"""
API do Super Conciliador V2 - Mercado Livre -> Conta Azul

VERSÃO MELHORADA com:
- Cruzamento de dados baseado no EXTRATO como fonte da verdade
- Detalhamento correto de receita, comissão e frete usando LIBERAÇÕES
- Validação cruzada dos valores
- Logging de transações não classificadas
- Tratamento correto de devoluções e chargebacks
- Suporte a arquivos ZIP com múltiplos CSVs (para períodos longos)

Endpoint:
    POST /conciliar - Recebe os relatórios CSV/ZIP e retorna um ZIP com os arquivos processados

Arquivos esperados (form-data) - Aceita CSV individual ou ZIP com múltiplos CSVs:
    - dinheiro: settlement report (obrigatório)
    - vendas: collection report (obrigatório)
    - pos_venda: after_collection report (obrigatório)
    - liberacoes: reserve-release report (obrigatório)
    - extrato: account_statement report (obrigatório)
    - retirada: withdraw report (opcional)

Fluxo de Cruzamento:
    1. EXTRATO é a fonte da verdade (movimentações reais na conta)
    2. LIBERAÇÕES detalha cada liberação (receita, taxas, frete)
    3. VENDAS enriquece com dados do pedido (produto, origem)
    4. DINHEIRO EM CONTA preenche previsões (ainda não liberadas)
    5. PÓS-VENDA adiciona contexto de devoluções
"""

import pandas as pd
import numpy as np
import os
import re
import io
import zipfile
import tempfile
import shutil
import logging
from datetime import datetime
from typing import Optional, Dict, List, Any, Tuple, Union
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font


# ==============================================================================
# FUNÇÕES PARA PROCESSAMENTO DE ZIP
# ==============================================================================

def is_zip_file(content: bytes) -> bool:
    """Verifica se o conteúdo é um arquivo ZIP pelo magic number"""
    return content[:4] == b'PK\x03\x04'


def extrair_csvs_do_zip(zip_content: bytes, skip_rows: int = 0, clean_json: bool = False) -> pd.DataFrame:
    """
    Extrai todos os arquivos CSV de um ZIP e concatena em um único DataFrame.

    Args:
        zip_content: Conteúdo binário do arquivo ZIP
        skip_rows: Número de linhas a pular no início de cada CSV
        clean_json: Se True, limpa campos JSON mal formatados

    Returns:
        DataFrame concatenado com todos os CSVs do ZIP
    """
    dataframes = []

    with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zip_file:
        # Listar todos os arquivos no ZIP
        csv_files = [f for f in zip_file.namelist()
                     if f.lower().endswith('.csv') and not f.startswith('__MACOSX')]

        if not csv_files:
            raise ValueError("Nenhum arquivo CSV encontrado dentro do ZIP")

        logger.info(f"ZIP contém {len(csv_files)} arquivo(s) CSV: {csv_files}")

        for csv_filename in csv_files:
            try:
                with zip_file.open(csv_filename) as csv_file:
                    content = csv_file.read()
                    content_str = content.decode('utf-8')

                    if clean_json:
                        # Remove campos JSON mal formatados
                        content_str = re.sub(r'"\{[^}]*(?:\{[^}]*\}[^}]*)*\}"', '""', content_str)

                    # Detectar separador automaticamente
                    lines = content_str.split('\n')
                    header_line = lines[skip_rows] if len(lines) > skip_rows else lines[0]
                    sep = ';' if header_line.count(';') > header_line.count(',') else ','

                    df = pd.read_csv(
                        io.StringIO(content_str),
                        sep=sep,
                        skiprows=skip_rows,
                        on_bad_lines='skip',
                        index_col=False
                    )

                    if not df.empty:
                        dataframes.append(df)
                        logger.info(f"  - {csv_filename}: {len(df)} linhas")

            except Exception as e:
                logger.warning(f"Erro ao processar {csv_filename} do ZIP: {str(e)}")
                continue

    if not dataframes:
        raise ValueError("Nenhum CSV válido foi extraído do ZIP")

    # Concatenar todos os DataFrames
    resultado = pd.concat(dataframes, ignore_index=True)
    logger.info(f"Total após concatenação: {len(resultado)} linhas")

    return resultado

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Super Conciliador API V2",
    description="API para conciliação de relatórios Mercado Livre com Conta Azul - Versão Melhorada",
    version="2.0.0"
)


# ==============================================================================
# CONSTANTES E CONFIGURAÇÃO
# ==============================================================================

# Plano de Contas (Conta Azul)
CA_CATS = {
    # RECEITAS (valores positivos)
    'RECEITA_ML': "1.1.1 MercadoLibre",
    'RECEITA_LOJA': "1.1.2 Loja Própria (E-commerce)",
    'RECEITA_BALCAO': "1.1.5 Vendas Diretas/Balcão",
    'ESTORNO_TAXA': "1.3.4 Descontos e Estornos de Taxas e Tarifas",
    'ESTORNO_FRETE': "1.3.7 Estorno de Frete sobre Vendas",

    # DESPESAS (valores negativos)
    'DEVOLUCAO': "1.2.1 Devoluções e Cancelamentos",
    'COMISSAO': "2.8.2 Comissões de Marketplace",
    'FRETE_ENVIO': "2.9.4 MercadoEnvios",
    'FRETE_REVERSO': "2.9.10 Logística Reversa",
    'DIFAL': "2.2.3 DIFAL (Diferencial de Alíquota)",
    'PAGAMENTO_CONTA': "2.1.1 Compra de Mercadorias",
    'MARKETING_ML': "2.7.3 Marketing em Marketplace",
    'OUTROS': "2.14.8 Despesas Eventuais",

    # NEUTRO
    'TRANSFERENCIA': "Transferências",
}

# ==============================================================================
# FUNÇÕES UTILITÁRIAS
# ==============================================================================

def clean_id(val) -> str:
    """Limpa IDs removendo .0 e espaços"""
    if pd.isna(val):
        return ""
    return str(val).replace('.0', '').strip()


def clean_float_extrato(val) -> float:
    """Converte valor do extrato (formato brasileiro) para float"""
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.replace('.', '').replace(',', '.')
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_float(val, default: float = 0.0) -> float:
    """Converte valor para float de forma segura"""
    if pd.isna(val):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def format_date(val) -> str:
    """Formata data para dd/mm/yyyy"""
    if pd.isna(val):
        return ""
    try:
        if isinstance(val, str):
            # Tenta parsear diferentes formatos
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                try:
                    return pd.to_datetime(val).strftime('%d/%m/%Y')
                except:
                    continue
        return pd.to_datetime(val).strftime('%d/%m/%Y')
    except:
        return ""


def processar_conciliacao(arquivos: Dict[str, pd.DataFrame], centro_custo: str = "NETAIR") -> Dict[str, Any]:
    """
    Processa a conciliação dos relatórios do Mercado Livre.

    NOVA LÓGICA V2:
    1. EXTRATO é a fonte da verdade (o que realmente movimentou)
    2. Para cada movimento do EXTRATO, busca detalhes no LIBERAÇÕES
    3. LIBERAÇÕES tem o breakdown: receita, comissão, frete
    4. VENDAS enriquece com dados do pedido
    5. DINHEIRO EM CONTA é usado apenas para PREVISÕES (não liberados ainda)

    Args:
        arquivos: Dicionário com DataFrames dos relatórios
        centro_custo: Centro de custo para os lançamentos (padrão: NETAIR)

    Returns:
        Dicionário com os DataFrames processados e estatísticas
    """

    dinheiro = arquivos['dinheiro']
    vendas = arquivos['vendas']
    pos_venda = arquivos['pos_venda']
    liberacoes = arquivos['liberacoes']
    extrato = arquivos['extrato']

    CENTRO_CUSTO = centro_custo

    # ==============================================================================
    # FASE 1: PREPARAÇÃO E INDEXAÇÃO DOS DADOS
    # ==============================================================================

    logger.info("Fase 1: Preparando e indexando dados...")

    # 1.1 Normalizar IDs em todos os DataFrames
    if 'SOURCE_ID' in dinheiro.columns:
        dinheiro['op_id'] = dinheiro['SOURCE_ID'].apply(clean_id)

    if 'Número da transação do Mercado Pago (operation_id)' in vendas.columns:
        vendas['op_id'] = vendas['Número da transação do Mercado Pago (operation_id)'].apply(clean_id)

    if 'ID da transação (operation_id)' in pos_venda.columns:
        pos_venda['op_id'] = pos_venda['ID da transação (operation_id)'].apply(clean_id)

    # 1.2 Criar mapa de ORIGEM da venda (ML, LOJA, BALCÃO)
    map_origem_venda = {}

    # Primeiro: se tem order_id do ML, é venda ML
    if 'Número da venda no Mercado Livre (order_id)' in vendas.columns:
        for _, row in vendas.iterrows():
            op_id = clean_id(row.get('Número da transação do Mercado Pago (operation_id)', ''))
            order_id = row.get('Número da venda no Mercado Livre (order_id)', '')
            if op_id and pd.notna(order_id) and str(order_id).strip() not in ['', 'nan']:
                map_origem_venda[op_id] = 'ML'

    # Segundo: verifica SUB_UNIT no dinheiro
    if 'SUB_UNIT' in dinheiro.columns:
        for _, row in dinheiro.iterrows():
            op_id = clean_id(row.get('SOURCE_ID', ''))
            sub_unit = str(row.get('SUB_UNIT', '')).lower()
            if op_id and op_id not in map_origem_venda:
                if 'point' in sub_unit:
                    map_origem_venda[op_id] = 'BALCAO'
                else:
                    map_origem_venda[op_id] = 'LOJA'

    def get_categoria_receita(op_id: str) -> str:
        """Retorna a categoria de receita baseada na origem da venda"""
        origem = map_origem_venda.get(op_id, 'LOJA')
        if origem == 'ML':
            return CA_CATS['RECEITA_ML']
        elif origem == 'BALCAO':
            return CA_CATS['RECEITA_BALCAO']
        else:
            return CA_CATS['RECEITA_LOJA']

    # 1.3 Criar mapas de dados das VENDAS para enriquecimento
    map_vendas = {}
    for _, row in vendas.iterrows():
        op_id = row.get('op_id', '')
        if op_id:
            map_vendas[op_id] = {
                'valor_produto': safe_float(row.get('Valor do produto (transaction_amount)', 0)),
                'frete_comprador': safe_float(row.get('Frete (shipping_cost)', 0)),
                'descricao': str(row.get('Descrição da operação (reason)', '')),
                'order_id': clean_id(row.get('Número da venda no Mercado Livre (order_id)', '')),
                'data_venda': row.get('Data da compra (date_created)', ''),
                'data_liberacao': row.get('Data de liberação do dinheiro (date_released)', ''),
                'status_envio': str(row.get('Status do envio (shipment_status)', '')),
            }

    # 1.4 Criar mapa do PÓS-VENDA para contexto de devoluções
    map_pos_venda = {}
    for _, row in pos_venda.iterrows():
        op_id = row.get('op_id', '')
        if op_id:
            map_pos_venda[op_id] = {
                'motivo': str(row.get('Motivo detalhado (reason_detail)', '')),
            }

    # ==============================================================================
    # FASE 2: INDEXAR LIBERAÇÕES POR SOURCE_ID E DESCRIPTION
    # ==============================================================================

    logger.info("Fase 2: Indexando liberações...")

    # Filtrar liberações válidas
    if 'RECORD_TYPE' in liberacoes.columns:
        liberacoes_filtrado = liberacoes[liberacoes['RECORD_TYPE'] != 'available_balance'].copy()
    elif 'SOURCE_ID' in liberacoes.columns:
        liberacoes_filtrado = liberacoes[liberacoes['SOURCE_ID'].notna()].copy()
    else:
        liberacoes_filtrado = liberacoes.copy()

    # Mapa de liberações por (SOURCE_ID, DESCRIPTION)
    # Estrutura: {op_id: {'payment': {...}, 'refund': [{...}, {...}], ...}}
    map_liberacoes = {}

    for _, row in liberacoes_filtrado.iterrows():
        op_id = clean_id(row.get('SOURCE_ID', ''))
        if not op_id:
            continue

        desc = str(row.get('DESCRIPTION', '')).lower().strip()

        # Extrair valores do LIBERAÇÕES
        dados = {
            'date': row.get('DATE', ''),
            'gross_amount': safe_float(row.get('GROSS_AMOUNT', 0)),
            'mp_fee': safe_float(row.get('MP_FEE_AMOUNT', 0)),
            'financing_fee': safe_float(row.get('FINANCING_FEE_AMOUNT', 0)),
            'shipping_fee': safe_float(row.get('SHIPPING_FEE_AMOUNT', 0)),
            'net_credit': safe_float(row.get('NET_CREDIT_AMOUNT', 0)),
            'net_debit': safe_float(row.get('NET_DEBIT_AMOUNT', 0)),
        }

        # Calcular valor líquido
        dados['net_amount'] = dados['net_credit'] - dados['net_debit']

        # Calcular comissão total (MP + parcelamento)
        dados['comissao_total'] = dados['mp_fee'] + dados['financing_fee']

        if op_id not in map_liberacoes:
            map_liberacoes[op_id] = {}

        # Armazenar por tipo de descrição
        # Alguns tipos podem ter múltiplos registros para o mesmo ID
        if desc in ['refund', 'chargeback', 'mediation', 'reserve_for_dispute']:
            if desc not in map_liberacoes[op_id]:
                map_liberacoes[op_id][desc] = []
            map_liberacoes[op_id][desc].append(dados)
        else:
            # Payment geralmente é único
            map_liberacoes[op_id][desc] = dados

    # ==============================================================================
    # FASE 3: IDENTIFICAR TRANSAÇÕES JÁ LIBERADAS (via LIBERAÇÕES)
    # ==============================================================================

    # IDs que já aparecem no LIBERAÇÕES = já foram processados
    ids_liberados = set(map_liberacoes.keys())
    logger.info(f"Total de IDs com liberação: {len(ids_liberados)}")

    # ==============================================================================
    # FASE 4: PROCESSAR EXTRATO (FONTE DA VERDADE)
    # ==============================================================================

    logger.info("Fase 4: Processando EXTRATO...")

    rows_conta_azul_confirmados = []
    rows_conta_azul_previsao = []
    rows_pagamento_conta = []
    rows_transferencias = []
    rows_nao_classificados = []  # Para rastreabilidade

    # Preparar EXTRATO
    extrato['Valor'] = extrato['TRANSACTION_NET_AMOUNT'].apply(clean_float_extrato)
    extrato['Data'] = pd.to_datetime(extrato['RELEASE_DATE'], dayfirst=True, errors='coerce')
    extrato['DataStr'] = extrato['Data'].dt.strftime('%d/%m/%Y')
    extrato['ID'] = extrato['REFERENCE_ID'].astype(str).str.replace('.0', '', regex=False).str.strip()

    def criar_lancamento(op_id: str, data_str: str, categoria: str, valor: float,
                         descricao: str, observacoes: str, centro: str = CENTRO_CUSTO) -> Dict:
        """Helper para criar lançamento padronizado"""
        return {
            'ID Operação': op_id,
            'Data de Competência': data_str,
            'Data de Pagamento': data_str,
            'Categoria': categoria,
            'Valor': round(valor, 2),
            'Centro de Custo': centro,
            'Descrição': descricao,
            'Observações': observacoes
        }

    def buscar_liberacao_por_tipo_e_valor(op_id: str, tipo_extrato: str, valor_extrato: float) -> Optional[Dict]:
        """
        Busca o registro correto no LIBERAÇÕES baseado no tipo de transação do extrato e valor.

        MAPEAMENTO EXTRATO -> LIBERAÇÕES:
        - "Liberação de dinheiro" -> payment
        - "Débito por dívida Reclamações..." -> mediation
        - "Reembolso..." -> refund
        - "Dinheiro retido..." -> reserve_for_dispute
        """
        if op_id not in map_liberacoes:
            return None

        tipo_lower = tipo_extrato.lower()
        lib_data = map_liberacoes[op_id]

        # Determinar qual DESCRIPTION buscar
        if 'liberação de dinheiro' in tipo_lower or 'liberacao de dinheiro' in tipo_lower:
            target_desc = 'payment'
        elif 'débito por dívida' in tipo_lower or 'debito por divida' in tipo_lower:
            target_desc = 'mediation'
        elif 'reembolso' in tipo_lower:
            target_desc = 'refund'
        elif 'dinheiro retido' in tipo_lower:
            target_desc = 'reserve_for_dispute'
        else:
            return None

        # Buscar o registro
        if target_desc in lib_data:
            dados = lib_data[target_desc]
            # Se for lista, buscar pelo valor
            if isinstance(dados, list):
                for item in dados:
                    if abs(item['net_amount'] - valor_extrato) < 0.10:
                        return item
                # Se não achou por valor exato, retorna o primeiro
                return dados[0] if dados else None
            else:
                # É um dict único
                return dados

        return None

    def detalhar_liberacao_payment(op_id: str, data_str: str, valor_extrato: float,
                                   descricao_base: str) -> List[Dict]:
        """
        Detalha uma liberação de pagamento usando dados do LIBERAÇÕES.

        LÓGICA CORRETA:
        - GROSS_AMOUNT = valor bruto (inclui frete do comprador)
        - SHIPPING_FEE_AMOUNT = frete (negativo = despesa do vendedor OU repasse do frete do comprador)
        - MP_FEE_AMOUNT = taxa do Mercado Pago (negativo = despesa)
        - FINANCING_FEE_AMOUNT = taxa de parcelamento (negativo = despesa)
        - NET_CREDIT - NET_DEBIT = valor líquido (deve bater com extrato)

        IMPORTANTE SOBRE FRETE:
        - O GROSS_AMOUNT inclui o frete pago pelo comprador
        - O SHIPPING_FEE_AMOUNT é NEGATIVO e representa o repasse desse frete ao ML
        - Para a RECEITA, usamos: GROSS_AMOUNT + SHIPPING_FEE (desconta o frete do comprador)
        - Isso faz a receita bater com o valor do produto mostrado no painel do ML

        IMPORTANTE: Um mesmo ID pode ter múltiplas entradas no LIBERAÇÕES
        (payment + refund parcial). Precisamos somar tudo para que bata com o extrato.
        """
        lancamentos = []

        # Buscar dados no mapa de liberações
        if op_id in map_liberacoes and 'payment' in map_liberacoes[op_id]:
            lib = map_liberacoes[op_id]['payment']

            # Valores do LIBERAÇÕES (payment)
            gross = lib['gross_amount']
            comissao = lib['comissao_total']  # MP + financing (já negativos)
            frete_lib = lib['shipping_fee']  # Negativo no LIBERAÇÕES
            liquido_calculado = lib['net_amount']

            # Verificar quem pagou o frete consultando VENDAS
            # VENDAS.Frete < 0  → VENDEDOR paga frete (despesa real)
            # VENDAS.Frete = 0  → COMPRADOR pagou (frete embutido no GROSS, é só repasse)
            frete_vendas = 0.0
            valor_produto = gross  # Default: usar GROSS
            if op_id in map_vendas:
                frete_vendas = map_vendas[op_id].get('frete_comprador', 0.0)
                valor_produto = map_vendas[op_id].get('valor_produto', gross)

            # Determinar se é frete do vendedor ou do comprador
            vendedor_paga_frete = frete_vendas < -0.01  # Negativo em VENDAS = vendedor paga

            if vendedor_paga_frete:
                # VENDEDOR paga frete: Receita = GROSS, Frete = despesa separada
                receita = gross
                frete_despesa = frete_lib  # Negativo, é despesa
            else:
                # COMPRADOR pagou frete: Receita = valor do produto (sem frete embutido)
                # O frete entra no GROSS e sai no SHIPPING_FEE, é só repasse
                receita = gross + frete_lib  # gross + (frete negativo) = valor produto
                frete_despesa = 0.0  # Não lançar frete como despesa (é repasse)

            # Verificar se há refund parcial associado
            refund_gross = 0.0
            refund_estorno_taxa = 0.0
            refund_estorno_frete = 0.0

            if 'refund' in map_liberacoes[op_id]:
                refunds = map_liberacoes[op_id]['refund']
                for ref in refunds:
                    refund_gross += ref['gross_amount']  # Negativo = devolvido
                    refund_estorno_taxa += ref['mp_fee'] + ref['financing_fee']  # Positivo = estornado
                    refund_estorno_frete += ref['shipping_fee']  # Positivo = estornado
                    liquido_calculado += ref['net_amount']

            # Validação: o líquido calculado deve bater com o extrato
            diferenca = abs(liquido_calculado - valor_extrato)
            if diferenca > 0.10:
                logger.warning(f"DISCREPÂNCIA op_id={op_id}: extrato={valor_extrato}, calculado={liquido_calculado}, diff={diferenca}")

            # LANÇAMENTO 1: Receita
            if abs(receita) > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    get_categoria_receita(op_id),
                    abs(receita),  # Receita sempre positiva
                    descricao_base,
                    "Receita de venda"
                ))

            # LANÇAMENTO 2: Comissão (sempre negativa)
            if abs(comissao) > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['COMISSAO'],
                    -abs(comissao),  # Despesa sempre negativa
                    descricao_base,
                    f"Tarifa ML (MP: {lib['mp_fee']:.2f} + Parc: {lib['financing_fee']:.2f})"
                ))

            # LANÇAMENTO 3: Frete (somente se VENDEDOR paga)
            if abs(frete_despesa) > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['FRETE_ENVIO'],
                    -abs(frete_despesa),  # Despesa sempre negativa
                    descricao_base,
                    "Frete de envio (MercadoEnvios)"
                ))

            # LANÇAMENTO 4: Refund parcial (se houver)
            if abs(refund_gross) > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['DEVOLUCAO'],
                    refund_gross,  # Já é negativo
                    descricao_base,
                    "Devolução parcial"
                ))

            # LANÇAMENTO 5: Estorno de taxa (se houve refund)
            if refund_estorno_taxa > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['ESTORNO_TAXA'],
                    refund_estorno_taxa,
                    descricao_base,
                    "Estorno de taxa (devolução parcial)"
                ))

            # LANÇAMENTO 6: Estorno de frete (se houve refund)
            if refund_estorno_frete > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['ESTORNO_FRETE'],
                    refund_estorno_frete,
                    descricao_base,
                    "Estorno de frete (devolução parcial)"
                ))

        else:
            # Fallback: não tem detalhes no LIBERAÇÕES
            # Tenta usar dados do DINHEIRO EM CONTA
            logger.info(f"op_id={op_id} sem detalhes em LIBERAÇÕES, usando fallback")

            if op_id in map_vendas:
                venda = map_vendas[op_id]
                receita = venda['valor_produto']
                if receita > 0:
                    # Calcula comissão por diferença
                    comissao = receita - valor_extrato
                    if comissao > 0:
                        lancamentos.append(criar_lancamento(
                            op_id, data_str,
                            get_categoria_receita(op_id),
                            receita,
                            descricao_base,
                            "Receita de venda (estimada)"
                        ))
                        lancamentos.append(criar_lancamento(
                            op_id, data_str,
                            CA_CATS['COMISSAO'],
                            -comissao,
                            descricao_base,
                            "Tarifa ML (calculada por diferença)"
                        ))
                    else:
                        # Não conseguiu calcular, usa valor total
                        lancamentos.append(criar_lancamento(
                            op_id, data_str,
                            get_categoria_receita(op_id),
                            valor_extrato,
                            descricao_base,
                            "Liberação de venda (sem detalhamento)"
                        ))
            else:
                # Último fallback: valor total como receita
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    get_categoria_receita(op_id),
                    valor_extrato,
                    descricao_base,
                    "Liberação de venda (sem detalhamento)"
                ))

        return lancamentos

    def detalhar_refund(op_id: str, data_str: str, valor_extrato: float,
                        descricao_base: str) -> List[Dict]:
        """
        Detalha um reembolso usando dados do LIBERAÇÕES.

        Em um refund:
        - GROSS_AMOUNT negativo = valor devolvido ao comprador
        - MP_FEE positivo = estorno da taxa (volta pro vendedor)
        - FINANCING_FEE positivo = estorno da taxa de parcelamento
        - SHIPPING_FEE positivo = estorno do frete
        """
        lancamentos = []

        if op_id in map_liberacoes and 'refund' in map_liberacoes[op_id]:
            refunds = map_liberacoes[op_id]['refund']
            # Pega o primeiro refund (ou soma se houver múltiplos)
            ref = refunds[0] if len(refunds) == 1 else refunds[0]  # TODO: somar múltiplos

            valor_devolvido = ref['gross_amount']  # Negativo
            estorno_mp_fee = ref['mp_fee']  # Positivo se estornado
            estorno_financing = ref['financing_fee']  # Positivo se estornado
            estorno_frete = ref['shipping_fee']  # Positivo se estornado

            # LANÇAMENTO 1: Devolução (valor devolvido ao comprador)
            if abs(valor_devolvido) > 0.01:
                # Se GROSS_AMOUNT é negativo, é devolução
                if valor_devolvido < 0:
                    lancamentos.append(criar_lancamento(
                        op_id, data_str,
                        CA_CATS['DEVOLUCAO'],
                        valor_devolvido,  # Já é negativo
                        descricao_base,
                        "Devolução de produto"
                    ))
                else:
                    # Se positivo, é estorno de devolução anterior
                    lancamentos.append(criar_lancamento(
                        op_id, data_str,
                        CA_CATS['ESTORNO_TAXA'],
                        valor_devolvido,
                        descricao_base,
                        "Estorno de devolução"
                    ))

            # LANÇAMENTO 2: Estorno de taxa MP (se positivo)
            if estorno_mp_fee > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['ESTORNO_TAXA'],
                    estorno_mp_fee,
                    descricao_base,
                    "Estorno taxa Mercado Livre"
                ))
            elif estorno_mp_fee < -0.01:
                # Taxa cobrada (raro em refund)
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['COMISSAO'],
                    estorno_mp_fee,
                    descricao_base,
                    "Taxa sobre devolução"
                ))

            # LANÇAMENTO 3: Estorno de taxa parcelamento
            if estorno_financing > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['ESTORNO_TAXA'],
                    estorno_financing,
                    descricao_base,
                    "Estorno taxa parcelamento"
                ))

            # LANÇAMENTO 4: Estorno de frete
            if estorno_frete > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['ESTORNO_FRETE'],
                    estorno_frete,
                    descricao_base,
                    "Estorno de frete"
                ))
            elif estorno_frete < -0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['FRETE_REVERSO'],
                    estorno_frete,
                    descricao_base,
                    "Frete de logística reversa"
                ))

        else:
            # Fallback: não tem detalhes
            if valor_extrato > 0:
                # Positivo = estorno
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['ESTORNO_TAXA'],
                    valor_extrato,
                    descricao_base,
                    "Estorno (sem detalhamento)"
                ))
            else:
                # Negativo = devolução
                lancamentos.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['DEVOLUCAO'],
                    valor_extrato,
                    descricao_base,
                    "Reembolso (sem detalhamento)"
                ))

        return lancamentos

    def detalhar_transacao_assertiva(op_id: str, tipo_extrato: str, data_str: str,
                                     valor_extrato: float, descricao_base: str) -> List[Dict]:
        """
        Detalha uma transação de forma assertiva usando o mapeamento correto entre
        EXTRATO e LIBERAÇÕES.

        Para IDs com múltiplas transações, busca o registro específico no LIBERAÇÕES
        que corresponde a esta linha do extrato.

        Retorna lista de lançamentos detalhados ou lista vazia se não conseguir detalhar.
        """
        lancamentos = []
        tipo_lower = tipo_extrato.lower()

        # Buscar registro correspondente no LIBERAÇÕES
        lib = buscar_liberacao_por_tipo_e_valor(op_id, tipo_extrato, valor_extrato)

        if not lib:
            return []  # Não encontrou, processar de forma simplificada

        # =========================================================================
        # LIBERAÇÃO DE DINHEIRO (payment) - Detalha receita, comissão, frete
        # =========================================================================
        if 'liberação de dinheiro' in tipo_lower or 'liberacao de dinheiro' in tipo_lower:
            gross = lib['gross_amount']
            comissao = lib['comissao_total']
            frete_lib = lib['shipping_fee']

            # Verificar quem pagou o frete consultando VENDAS
            # VENDAS.Frete < 0  → VENDEDOR paga frete (despesa real)
            # VENDAS.Frete = 0  → COMPRADOR pagou (frete embutido no GROSS, é só repasse)
            frete_vendas = 0.0
            if op_id in map_vendas:
                frete_vendas = map_vendas[op_id].get('frete_comprador', 0.0)

            vendedor_paga_frete = frete_vendas < -0.01

            if vendedor_paga_frete:
                # VENDEDOR paga frete: Receita = GROSS, Frete = despesa separada
                receita = gross
                frete_despesa = frete_lib
            else:
                # COMPRADOR pagou frete: Receita = valor do produto (sem frete)
                receita = gross + frete_lib  # gross + (frete negativo) = valor produto
                frete_despesa = 0.0

            if abs(receita) > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str, get_categoria_receita(op_id),
                    abs(receita), descricao_base, "Receita de venda"
                ))

            if abs(comissao) > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str, CA_CATS['COMISSAO'],
                    -abs(comissao), descricao_base,
                    f"Tarifa ML (MP: {lib['mp_fee']:.2f} + Parc: {lib['financing_fee']:.2f})"
                ))

            # Frete somente se VENDEDOR paga
            if abs(frete_despesa) > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str, CA_CATS['FRETE_ENVIO'],
                    -abs(frete_despesa), descricao_base, "Frete de envio (MercadoEnvios)"
                ))

        # =========================================================================
        # DÉBITO POR DÍVIDA / MEDIAÇÃO - Valor negativo direto
        # =========================================================================
        elif 'débito por dívida' in tipo_lower or 'debito por divida' in tipo_lower:
            # mediation: GROSS_AMOUNT é o valor da dívida (negativo)
            # Não tem detalhamento, é um valor direto
            lancamentos.append(criar_lancamento(
                op_id, data_str, CA_CATS['DEVOLUCAO'],
                valor_extrato, descricao_base, "Débito por reclamação/mediação ML"
            ))

        # =========================================================================
        # REEMBOLSO (refund) - Detalha estornos de taxa e frete
        # =========================================================================
        elif 'reembolso' in tipo_lower:
            # refund: MP_FEE e SHIPPING_FEE são positivos = estornados
            estorno_taxa = lib['mp_fee'] + lib['financing_fee']
            estorno_frete = lib['shipping_fee']

            # Se gross_amount for negativo, é devolução de valor ao comprador
            if lib['gross_amount'] < -0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str, CA_CATS['DEVOLUCAO'],
                    lib['gross_amount'], descricao_base, "Devolução ao comprador"
                ))

            # Estorno de taxa (positivo)
            if estorno_taxa > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str, CA_CATS['ESTORNO_TAXA'],
                    estorno_taxa, descricao_base, "Estorno de taxa ML"
                ))

            # Estorno de frete (positivo)
            if estorno_frete > 0.01:
                lancamentos.append(criar_lancamento(
                    op_id, data_str, CA_CATS['ESTORNO_FRETE'],
                    estorno_frete, descricao_base, "Estorno de frete"
                ))

        # =========================================================================
        # DINHEIRO RETIDO (reserve_for_dispute)
        # =========================================================================
        elif 'dinheiro retido' in tipo_lower:
            if valor_extrato < 0:
                lancamentos.append(criar_lancamento(
                    op_id, data_str, CA_CATS['DEVOLUCAO'],
                    valor_extrato, descricao_base, "Dinheiro retido (bloqueio por disputa)"
                ))
            else:
                lancamentos.append(criar_lancamento(
                    op_id, data_str, CA_CATS['ESTORNO_TAXA'],
                    valor_extrato, descricao_base, "Dinheiro liberado (desbloqueio)"
                ))

        return lancamentos

    # ==============================================================================
    # FASE 5: PROCESSAR EXTRATO LINHA POR LINHA
    # ==============================================================================

    logger.info("Fase 5: Processando cada linha do EXTRATO...")

    # Identificar quais IDs têm múltiplas transações no extrato
    ids_multiplos = extrato.groupby('ID').size()
    ids_multiplos = set(ids_multiplos[ids_multiplos > 1].index)
    logger.info(f"IDs com múltiplas transações no extrato: {len(ids_multiplos)}")

    for idx, row in extrato.iterrows():
        try:
            op_id = row['ID']
            tipo_transacao = str(row.get('TRANSACTION_TYPE', ''))
            val = row['Valor']
            data_str = row['DataStr']
            tipo_lower = tipo_transacao.lower()

            # Ignorar valores zerados
            if abs(val) < 0.01:
                continue

            descricao_base = f"{op_id} - {tipo_transacao[:50]}"

            # =====================================================================
            # CATEGORIA 1: TRANSFERÊNCIAS (PIX, TED, etc.)
            # =====================================================================
            if 'ransfer' in tipo_lower:
                is_pix_recebido = 'pix recebid' in tipo_lower
                is_interno = any(x in tipo_lower for x in ['netparts', 'jonathan', 'netair'])

                if is_pix_recebido and not is_interno and val > 0:
                    # PIX recebido de cliente = venda
                    rows_conta_azul_confirmados.append(criar_lancamento(
                        op_id, data_str,
                        get_categoria_receita(op_id),
                        val,
                        descricao_base,
                        "PIX recebido (venda externa)"
                    ))
                else:
                    # Transferência normal
                    rows_transferencias.append(criar_lancamento(
                        op_id, data_str,
                        CA_CATS['TRANSFERENCIA'],
                        val,
                        descricao_base,
                        tipo_transacao,
                        centro=""
                    ))
                continue

            # =====================================================================
            # CATEGORIA 2: LIBERAÇÃO DE DINHEIRO CANCELADA
            # =====================================================================
            if 'liberação de dinheiro cancelada' in tipo_lower or 'liberacao de dinheiro cancelada' in tipo_lower:
                if val > 0:
                    rows_conta_azul_confirmados.append(criar_lancamento(
                        op_id, data_str, CA_CATS['ESTORNO_TAXA'], val,
                        descricao_base, "Estorno de liberação cancelada"
                    ))
                else:
                    rows_conta_azul_confirmados.append(criar_lancamento(
                        op_id, data_str, CA_CATS['DEVOLUCAO'], val,
                        descricao_base, "Liberação cancelada (chargeback)"
                    ))
                continue

            # =====================================================================
            # CATEGORIA 3: PAGAMENTO FATURA CARTÃO MP (vai para transferências)
            # =====================================================================
            if 'pagamento' in tipo_lower and 'cartão de crédito' in tipo_lower:
                rows_transferencias.append(criar_lancamento(
                    op_id, data_str,
                    CA_CATS['TRANSFERENCIA'],
                    val,
                    descricao_base,
                    "Pagamento fatura cartão Mercado Pago",
                    centro=""
                ))
                continue

            # =====================================================================
            # CATEGORIA 4: LIBERAÇÃO DE DINHEIRO (VENDA)
            # Esta é a categoria principal - detalha usando LIBERAÇÕES
            # =====================================================================
            is_liberacao = 'liberação de dinheiro' in tipo_lower or 'liberacao de dinheiro' in tipo_lower
            has_payment_data = op_id in map_liberacoes and 'payment' in map_liberacoes[op_id]
            is_id_multiplo = op_id in ids_multiplos

            if is_liberacao and has_payment_data:
                if is_id_multiplo:
                    # Para IDs múltiplos, usar detalhamento assertivo
                    lancamentos = detalhar_transacao_assertiva(op_id, tipo_transacao, data_str, val, descricao_base)
                    if lancamentos:
                        rows_conta_azul_confirmados.extend(lancamentos)
                    else:
                        # Fallback: valor direto
                        rows_conta_azul_confirmados.append(criar_lancamento(
                            op_id, data_str, get_categoria_receita(op_id), val,
                            descricao_base, "Liberação de venda"
                        ))
                else:
                    # Para IDs únicos, usar detalhamento completo (com refund parcial se houver)
                    lancamentos = detalhar_liberacao_payment(op_id, data_str, val, descricao_base)
                    rows_conta_azul_confirmados.extend(lancamentos)
                continue
            elif is_liberacao:
                # Liberação sem dados detalhados - usa valor do extrato
                rows_conta_azul_confirmados.append(criar_lancamento(
                    op_id, data_str,
                    get_categoria_receita(op_id),
                    val,
                    descricao_base,
                    "Liberação de venda"
                ))
                continue

            # =====================================================================
            # CATEGORIA 5: REEMBOLSO
            # Detalha usando dados do LIBERAÇÕES
            # =====================================================================
            if tipo_transacao.strip().startswith('Reembolso') or 'reembolso' in tipo_lower:
                if is_id_multiplo:
                    # Para IDs múltiplos, usar detalhamento assertivo
                    lancamentos = detalhar_transacao_assertiva(op_id, tipo_transacao, data_str, val, descricao_base)
                    if lancamentos:
                        rows_conta_azul_confirmados.extend(lancamentos)
                    else:
                        # Fallback: valor direto
                        if val > 0:
                            rows_conta_azul_confirmados.append(criar_lancamento(
                                op_id, data_str, CA_CATS['ESTORNO_TAXA'], val,
                                descricao_base, "Estorno (envío cancelado/devolução)"
                            ))
                        else:
                            rows_conta_azul_confirmados.append(criar_lancamento(
                                op_id, data_str, CA_CATS['DEVOLUCAO'], val,
                                descricao_base, "Devolução ao comprador"
                            ))
                else:
                    # Para IDs únicos, detalhar usando LIBERAÇÕES
                    lancamentos = detalhar_refund(op_id, data_str, val, descricao_base)
                    rows_conta_azul_confirmados.extend(lancamentos)
                continue

            # =====================================================================
            # CATEGORIA 6: DINHEIRO RETIDO (Disputa em andamento)
            # =====================================================================
            if 'dinheiro retido' in tipo_lower:
                # Dinheiro retido = bloqueio temporário por disputa
                # Valor negativo = bloqueou, valor positivo = desbloqueou
                if val < 0:
                    rows_conta_azul_confirmados.append(criar_lancamento(
                        op_id, data_str, CA_CATS['DEVOLUCAO'], val,
                        descricao_base, "Dinheiro retido (bloqueio por disputa)"
                    ))
                else:
                    rows_conta_azul_confirmados.append(criar_lancamento(
                        op_id, data_str, CA_CATS['ESTORNO_TAXA'], val,
                        descricao_base, "Dinheiro liberado (desbloqueio)"
                    ))
                continue

            # =====================================================================
            # CATEGORIA 7: OUTRAS TRANSAÇÕES ESPECÍFICAS
            # =====================================================================

            # DIFAL / Impostos
            if 'difal' in tipo_lower or 'imposto interestadual' in tipo_lower or 'aliquota' in tipo_lower:
                rows_conta_azul_confirmados.append(criar_lancamento(
                    op_id, data_str, CA_CATS['DIFAL'], val,
                    descricao_base, "DIFAL/Imposto Interestadual"
                ))
                continue

            # Pagamento de contas
            if 'pagamento de contas' in tipo_lower:
                rows_pagamento_conta.append(criar_lancamento(
                    op_id, data_str, CA_CATS['PAGAMENTO_CONTA'], val,
                    descricao_base, "Pagamento de conta via MP"
                ))
                continue

            # Pagamento/QR (PIX enviado ou recebido)
            if 'pagamento' in tipo_lower or 'qr' in tipo_lower:
                if val < 0:
                    rows_pagamento_conta.append(criar_lancamento(
                        op_id, data_str, CA_CATS['PAGAMENTO_CONTA'], val,
                        descricao_base, "Pagamento enviado via PIX/QR"
                    ))
                else:
                    rows_conta_azul_confirmados.append(criar_lancamento(
                        op_id, data_str, get_categoria_receita(op_id), val,
                        descricao_base, "Pagamento recebido via PIX/QR"
                    ))
                continue

            # Entrada de dinheiro
            if 'entrada' in tipo_lower:
                rows_conta_azul_confirmados.append(criar_lancamento(
                    op_id, data_str, get_categoria_receita(op_id), val,
                    descricao_base, "Entrada de dinheiro"
                ))
                continue

            # Débitos diversos
            if 'débito' in tipo_lower or 'debito' in tipo_lower or 'dívida' in tipo_lower or 'divida' in tipo_lower:
                # Para IDs múltiplos com reclamação, tentar detalhamento assertivo
                if is_id_multiplo and 'reclama' in tipo_lower:
                    lancamentos = detalhar_transacao_assertiva(op_id, tipo_transacao, data_str, val, descricao_base)
                    if lancamentos:
                        rows_conta_azul_confirmados.extend(lancamentos)
                        continue

                # Fallback: categorizar pelo tipo
                if 'reclama' in tipo_lower:
                    categoria = CA_CATS['DEVOLUCAO']
                    obs = "Débito por reclamação ML"
                elif 'envio' in tipo_lower:
                    categoria = CA_CATS['FRETE_ENVIO']
                    obs = "Débito de envio"
                elif 'troca' in tipo_lower:
                    categoria = CA_CATS['DEVOLUCAO']
                    obs = "Débito por troca de produto"
                elif 'fatura' in tipo_lower:
                    categoria = CA_CATS['MARKETING_ML']
                    obs = "Product ADS"
                elif 'retido' in tipo_lower:
                    categoria = CA_CATS['DEVOLUCAO']
                    obs = "Dinheiro retido por disputa"
                else:
                    categoria = CA_CATS['OUTROS']
                    obs = "Débito/Dívida ML"

                rows_conta_azul_confirmados.append(criar_lancamento(
                    op_id, data_str, categoria, val, descricao_base, obs
                ))
                continue

            # Bônus de envio
            if 'bônus' in tipo_lower or 'bonus' in tipo_lower:
                rows_conta_azul_confirmados.append(criar_lancamento(
                    op_id, data_str, CA_CATS['ESTORNO_FRETE'], val,
                    descricao_base, "Bônus de envio"
                ))
                continue

            # Compra no ML
            if 'compra' in tipo_lower:
                rows_pagamento_conta.append(criar_lancamento(
                    op_id, data_str, CA_CATS['PAGAMENTO_CONTA'], val,
                    descricao_base, "Compra no Mercado Livre"
                ))
                continue

            # =====================================================================
            # CATEGORIA 8: NÃO CLASSIFICADO (para revisão)
            # =====================================================================
            rows_nao_classificados.append({
                'op_id': op_id,
                'tipo': tipo_transacao,
                'valor': val,
                'data': data_str
            })

            rows_conta_azul_confirmados.append(criar_lancamento(
                op_id, data_str, CA_CATS['OUTROS'], val,
                descricao_base, f"REVISAR: {tipo_transacao[:30]}"
            ))

        except Exception as e:
            logger.error(f"Erro processando linha {idx}: {str(e)}")
            continue

    logger.info(f"Processadas {len(rows_conta_azul_confirmados)} transações confirmadas")
    logger.info(f"Transações não classificadas: {len(rows_nao_classificados)}")

    # ==============================================================================
    # FASE 6: PROCESSAR PREVISÕES (DINHEIRO EM CONTA não liberado)
    # ==============================================================================

    logger.info("Fase 6: Processando PREVISÕES (dinheiro não liberado)...")

    for _, row in dinheiro.iterrows():
        try:
            op_id = row.get('op_id', '')
            if not op_id:
                continue

            tipo_op = str(row.get('TRANSACTION_TYPE', ''))

            # Se já foi liberado (está no mapa de liberações), pula
            if op_id in map_liberacoes:
                continue

            # Extrair datas
            data_competencia = format_date(row.get('TRANSACTION_DATE', ''))
            if op_id in map_vendas:
                data_venda = map_vendas[op_id].get('data_venda', '')
                if data_venda:
                    data_competencia = format_date(data_venda)

            data_caixa = format_date(row.get('MONEY_RELEASE_DATE', ''))
            if not data_caixa and op_id in map_vendas:
                data_caixa = format_date(map_vendas[op_id].get('data_liberacao', ''))

            # Descrição
            id_pedido = clean_id(row.get('EXTERNAL_REFERENCE', ''))
            if not id_pedido:
                id_pedido = clean_id(row.get('ORDER_ID', ''))
            desc_part = f"Pedido {id_pedido}" if id_pedido else f"Op {op_id}"
            descricao_base = f"{op_id} - {desc_part}"

            if tipo_op == 'SETTLEMENT':
                # Obter valores
                if op_id in map_vendas:
                    val_receita = map_vendas[op_id]['valor_produto']
                else:
                    val_receita = safe_float(row.get('TRANSACTION_AMOUNT', 0))

                val_liquido = safe_float(row.get('REAL_AMOUNT', 0))
                val_frete = safe_float(row.get('SHIPPING_FEE_AMOUNT', 0))

                # Se valor negativo, é pagamento - vai para PREVISÃO (não confirmado)
                # Esses pagamentos só devem aparecer nos CONFIRMADOS quando estiverem no EXTRATO
                if val_receita < 0:
                    rows_conta_azul_previsao.append(criar_lancamento(
                        op_id, data_caixa or data_competencia,
                        CA_CATS['PAGAMENTO_CONTA'], val_receita,
                        descricao_base, "Pagamento via Mercado Pago (PREVISÃO)"
                    ))
                    continue

                # Receita (PREVISÃO)
                rows_conta_azul_previsao.append(criar_lancamento(
                    op_id, data_competencia,
                    get_categoria_receita(op_id), val_receita,
                    descricao_base, "Receita de venda (PREVISÃO)"
                ))

                # Calcular comissão
                if val_frete > 0:
                    val_frete = -val_frete
                val_comissao = round(val_receita + val_frete - val_liquido, 2)

                if abs(val_comissao) > 0.01:
                    rows_conta_azul_previsao.append(criar_lancamento(
                        op_id, data_competencia,
                        CA_CATS['COMISSAO'], -abs(val_comissao),
                        descricao_base, "Tarifa ML (PREVISÃO)"
                    ))

                if val_frete != 0:
                    rows_conta_azul_previsao.append(criar_lancamento(
                        op_id, data_competencia,
                        CA_CATS['FRETE_ENVIO'], val_frete,
                        descricao_base, "Frete (PREVISÃO)"
                    ))

            elif tipo_op in ['CHARGEBACK', 'REFUND', 'CANCELLATION', 'DISPUTE']:
                val = safe_float(row.get('TRANSACTION_AMOUNT', 0))
                if val > 0:
                    val = -val  # Devoluções são negativas

                rows_conta_azul_previsao.append(criar_lancamento(
                    op_id, data_competencia,
                    CA_CATS['DEVOLUCAO'], val,
                    descricao_base, f"{tipo_op} (PREVISÃO)"
                ))

            elif tipo_op in ['PAYOUT', 'MONEY_TRANSFER'] or 'RETIRADA' in tipo_op.upper():
                # Ignorar saques e transferências
                continue

            else:
                # Outros tipos
                val = safe_float(row.get('REAL_AMOUNT', 0))
                if abs(val) > 0.01:
                    rows_conta_azul_previsao.append(criar_lancamento(
                        op_id, data_competencia,
                        CA_CATS['OUTROS'], val,
                        descricao_base, f"REVISAR: {tipo_op} (PREVISÃO)",
                        centro=""
                    ))

        except Exception as e:
            logger.error(f"Erro processando previsão op_id={op_id}: {str(e)}")
            continue

    logger.info(f"Processadas {len(rows_conta_azul_previsao)} previsões")

    # ==============================================================================
    # FASE 7: ESTATÍSTICAS E RETORNO
    # ==============================================================================

    # Estatísticas de origem
    origens_count = {'ML': 0, 'LOJA': 0, 'BALCAO': 0}
    for origem in map_origem_venda.values():
        origens_count[origem] = origens_count.get(origem, 0) + 1

    # Log de transações não classificadas para debug
    if rows_nao_classificados:
        logger.warning(f"⚠️ {len(rows_nao_classificados)} transações não classificadas:")
        for nc in rows_nao_classificados[:10]:  # Mostrar até 10
            logger.warning(f"  - {nc['op_id']}: {nc['tipo']} = R$ {nc['valor']:.2f}")

    return {
        'confirmados': rows_conta_azul_confirmados,
        'previsao': rows_conta_azul_previsao,
        'pagamentos': rows_pagamento_conta,
        'transferencias': rows_transferencias,
        'nao_classificados': rows_nao_classificados,
        'stats': {
            'confirmados': len(rows_conta_azul_confirmados),
            'previsao': len(rows_conta_azul_previsao),
            'pagamentos': len(rows_pagamento_conta),
            'transferencias': len(rows_transferencias),
            'nao_classificados': len(rows_nao_classificados),
            'origens': origens_count,
            'ids_com_liberacao': len(ids_liberados)
        }
    }


def gerar_csv_conta_azul(rows: List[Dict], output_path: str) -> bool:
    """Gera arquivo CSV no formato Conta Azul"""
    if not rows:
        return False

    df = pd.DataFrame(rows)

    if df.empty:
        return False

    df['Valor'] = df['Valor'].round(2)
    df = df[df['Valor'] != 0]
    df['Data de Vencimento'] = df['Data de Pagamento']

    df['Cliente/Fornecedor'] = "MERCADO LIVRE"
    df['CNPJ/CPF Cliente/Fornecedor'] = "03007331000141"

    cols = ['Data de Competência', 'Data de Vencimento', 'Data de Pagamento', 'Valor',
            'Categoria', 'Descrição', 'Cliente/Fornecedor', 'CNPJ/CPF Cliente/Fornecedor',
            'Centro de Custo', 'Observações']

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    df[cols].to_csv(output_path, index=False, sep=';', encoding='utf-8-sig')
    return True


def gerar_xlsx_completo(rows: List[Dict], output_path: str) -> bool:
    """Gera arquivo XLSX com todas as transações"""
    if not rows:
        return False

    df = pd.DataFrame(rows)

    if df.empty:
        return False

    df['Valor'] = df['Valor'].round(2)
    df = df[df['Valor'] != 0]
    df['Data de Vencimento'] = df['Data de Pagamento']

    df['Cliente/Fornecedor'] = "MERCADO LIVRE"
    df['CNPJ/CPF Cliente/Fornecedor'] = "03007331000141"

    cols = ['Data de Competência', 'Data de Vencimento', 'Data de Pagamento', 'Valor',
            'Categoria', 'Descrição', 'Cliente/Fornecedor', 'CNPJ/CPF Cliente/Fornecedor',
            'Centro de Custo', 'Observações']

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Importação Conta Azul"

    header_font = Font(bold=True)

    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, row_data in enumerate(df[cols].values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    return True


def gerar_xlsx_resumo(rows: List[Dict], output_path: str) -> bool:
    """Gera arquivo XLSX com dados agrupados por Data de Pagamento e Categoria"""
    if not rows:
        return False

    df = pd.DataFrame(rows)

    if df.empty:
        return False

    df['Valor'] = df['Valor'].round(2)
    df = df[df['Valor'] != 0]
    df['Data de Vencimento'] = df['Data de Pagamento']

    df['Cliente/Fornecedor'] = "MERCADO LIVRE"
    df['CNPJ/CPF Cliente/Fornecedor'] = "03007331000141"

    df_grouped = df.groupby(['Data de Pagamento', 'Categoria'], as_index=False).agg({
        'Data de Competência': 'first',
        'Data de Vencimento': 'first',
        'Valor': 'sum',
        'Descrição': lambda x: f"Resumo {len(x)} transações",
        'Cliente/Fornecedor': 'first',
        'CNPJ/CPF Cliente/Fornecedor': 'first',
        'Centro de Custo': 'first',
        'Observações': lambda x: f"{len(x)} lançamentos agrupados"
    })

    df_grouped = df_grouped.sort_values('Data de Pagamento')

    cols = ['Data de Competência', 'Data de Vencimento', 'Data de Pagamento', 'Valor',
            'Categoria', 'Descrição', 'Cliente/Fornecedor', 'CNPJ/CPF Cliente/Fornecedor',
            'Centro de Custo', 'Observações']

    wb = Workbook()
    ws = wb.active
    ws.title = "Importação Conta Azul"

    header_font = Font(bold=True)

    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, row_data in enumerate(df_grouped[cols].values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    return True


# ==============================================================================
# ENDPOINTS DA API
# ==============================================================================

@app.get("/")
async def root():
    """Endpoint de health check"""
    return {
        "status": "online",
        "service": "Super Conciliador API",
        "version": "1.0.0"
    }


@app.post("/conciliar")
async def conciliar(
    dinheiro: UploadFile = File(..., description="Arquivo settlement (dinheiro em conta) - CSV ou ZIP"),
    vendas: UploadFile = File(..., description="Arquivo collection (vendas) - CSV ou ZIP"),
    pos_venda: UploadFile = File(..., description="Arquivo after_collection (pós venda) - CSV ou ZIP"),
    liberacoes: UploadFile = File(..., description="Arquivo reserve-release (liberações) - CSV ou ZIP"),
    extrato: UploadFile = File(..., description="Arquivo account_statement (extrato) - CSV ou ZIP"),
    retirada: Optional[UploadFile] = File(None, description="Arquivo withdraw (retirada) - opcional - CSV ou ZIP"),
    centro_custo: str = Form("NETAIR", description="Centro de custo para os lançamentos")
):
    """
    Processa os relatórios do Mercado Livre e retorna um ZIP com os arquivos de importação.

    ## Arquivos de entrada (CSV ou ZIP):
    Cada campo aceita um arquivo CSV individual OU um arquivo ZIP contendo múltiplos CSVs.
    Quando um ZIP é enviado, todos os CSVs dentro dele são extraídos e concatenados automaticamente.
    Isso é útil para períodos longos que geram múltiplos arquivos compactados.

    - **dinheiro**: settlement report (obrigatório) - CSV ou ZIP
    - **vendas**: collection report (obrigatório) - CSV ou ZIP
    - **pos_venda**: after_collection report (obrigatório) - CSV ou ZIP
    - **liberacoes**: reserve-release report (obrigatório) - CSV ou ZIP
    - **extrato**: account_statement report (obrigatório) - CSV ou ZIP
    - **retirada**: withdraw report (opcional) - CSV ou ZIP

    ## Parâmetros adicionais:
    - **centro_custo**: Centro de custo para os lançamentos (padrão: NETAIR)

    ## Arquivos de saída (ZIP):
    - IMPORTACAO_CONTA_AZUL_CONFIRMADOS.csv
    - IMPORTACAO_CONTA_AZUL_CONFIRMADOS.xlsx
    - IMPORTACAO_CONTA_AZUL_CONFIRMADOS_RESUMO.xlsx
    - IMPORTACAO_CONTA_AZUL_PREVISAO.csv
    - IMPORTACAO_CONTA_AZUL_PREVISAO.xlsx
    - IMPORTACAO_CONTA_AZUL_PREVISAO_RESUMO.xlsx
    - PAGAMENTO_CONTAS.csv
    - PAGAMENTO_CONTAS.xlsx
    - TRANSFERENCIAS.csv
    - TRANSFERENCIAS.xlsx
    """

    temp_dir = tempfile.mkdtemp()

    try:
        # Carregar DataFrames dos arquivos enviados
        arquivos = {}

        # Função auxiliar para ler CSV ou ZIP com detecção automática
        async def ler_csv(upload_file: UploadFile, key: str, skip_rows: int = 0, clean_json: bool = False):
            """
            Lê um arquivo CSV ou ZIP contendo múltiplos CSVs.

            Se o arquivo for um ZIP, extrai todos os CSVs e concatena em um único DataFrame.
            Isso é útil quando períodos longos geram múltiplos arquivos compactados.
            """
            content = await upload_file.read()

            # Verificar se é um arquivo ZIP
            if is_zip_file(content):
                logger.info(f"Arquivo '{key}' detectado como ZIP - extraindo e concatenando CSVs...")
                return extrair_csvs_do_zip(content, skip_rows=skip_rows, clean_json=clean_json)

            # Processar como CSV normal
            content_str = content.decode('utf-8')

            if clean_json:
                # Remove campos JSON mal formatados (METADATA com aspas internas não escapadas)
                # Pattern captura desde "{ até }" incluindo JSON aninhado
                content_str = re.sub(r'"\{[^}]*(?:\{[^}]*\}[^}]*)*\}"', '""', content_str)

            # Detectar separador automaticamente (verifica primeira linha após skip_rows)
            lines = content_str.split('\n')
            header_line = lines[skip_rows] if len(lines) > skip_rows else lines[0]

            # Conta ocorrências de ; e , na linha de cabeçalho (fora de aspas)
            sep = ';' if header_line.count(';') > header_line.count(',') else ','

            return pd.read_csv(
                io.StringIO(content_str),
                sep=sep,
                skiprows=skip_rows,
                on_bad_lines='skip',
                index_col=False
            )

        # Carregar arquivos obrigatórios
        try:
            arquivos['dinheiro'] = await ler_csv(dinheiro, 'dinheiro', clean_json=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo 'dinheiro': {str(e)}")

        try:
            arquivos['vendas'] = await ler_csv(vendas, 'vendas')
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo 'vendas': {str(e)}")

        try:
            arquivos['pos_venda'] = await ler_csv(pos_venda, 'pos_venda')
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo 'pos_venda': {str(e)}")

        try:
            arquivos['liberacoes'] = await ler_csv(liberacoes, 'liberacoes', clean_json=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo 'liberacoes': {str(e)}")

        try:
            arquivos['extrato'] = await ler_csv(extrato, 'extrato', skip_rows=3)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo 'extrato': {str(e)}")

        # Arquivo opcional
        if retirada:
            try:
                arquivos['retirada'] = await ler_csv(retirada, 'retirada')
            except:
                arquivos['retirada'] = pd.DataFrame()
        else:
            arquivos['retirada'] = pd.DataFrame()

        # Processar conciliação
        try:
            resultado = processar_conciliacao(arquivos, centro_custo=centro_custo)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao processar conciliação: {str(e)}")

        # Gerar arquivos de saída
        arquivos_gerados = []

        # CSVs
        if gerar_csv_conta_azul(resultado['confirmados'], os.path.join(temp_dir, 'IMPORTACAO_CONTA_AZUL_CONFIRMADOS.csv')):
            arquivos_gerados.append('IMPORTACAO_CONTA_AZUL_CONFIRMADOS.csv')

        if gerar_csv_conta_azul(resultado['previsao'], os.path.join(temp_dir, 'IMPORTACAO_CONTA_AZUL_PREVISAO.csv')):
            arquivos_gerados.append('IMPORTACAO_CONTA_AZUL_PREVISAO.csv')

        if gerar_csv_conta_azul(resultado['pagamentos'], os.path.join(temp_dir, 'PAGAMENTO_CONTAS.csv')):
            arquivos_gerados.append('PAGAMENTO_CONTAS.csv')

        if gerar_csv_conta_azul(resultado['transferencias'], os.path.join(temp_dir, 'TRANSFERENCIAS.csv')):
            arquivos_gerados.append('TRANSFERENCIAS.csv')

        # XLSXs Completos
        if gerar_xlsx_completo(resultado['confirmados'], os.path.join(temp_dir, 'IMPORTACAO_CONTA_AZUL_CONFIRMADOS.xlsx')):
            arquivos_gerados.append('IMPORTACAO_CONTA_AZUL_CONFIRMADOS.xlsx')

        if gerar_xlsx_completo(resultado['previsao'], os.path.join(temp_dir, 'IMPORTACAO_CONTA_AZUL_PREVISAO.xlsx')):
            arquivos_gerados.append('IMPORTACAO_CONTA_AZUL_PREVISAO.xlsx')

        if gerar_xlsx_completo(resultado['pagamentos'], os.path.join(temp_dir, 'PAGAMENTO_CONTAS.xlsx')):
            arquivos_gerados.append('PAGAMENTO_CONTAS.xlsx')

        if gerar_xlsx_completo(resultado['transferencias'], os.path.join(temp_dir, 'TRANSFERENCIAS.xlsx')):
            arquivos_gerados.append('TRANSFERENCIAS.xlsx')

        # XLSXs Resumidos
        if gerar_xlsx_resumo(resultado['confirmados'], os.path.join(temp_dir, 'IMPORTACAO_CONTA_AZUL_CONFIRMADOS_RESUMO.xlsx')):
            arquivos_gerados.append('IMPORTACAO_CONTA_AZUL_CONFIRMADOS_RESUMO.xlsx')

        if gerar_xlsx_resumo(resultado['previsao'], os.path.join(temp_dir, 'IMPORTACAO_CONTA_AZUL_PREVISAO_RESUMO.xlsx')):
            arquivos_gerados.append('IMPORTACAO_CONTA_AZUL_PREVISAO_RESUMO.xlsx')

        if not arquivos_gerados:
            raise HTTPException(status_code=500, detail="Nenhum arquivo foi gerado. Verifique os dados de entrada.")

        # Criar ZIP em memória
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for arquivo in arquivos_gerados:
                file_path = os.path.join(temp_dir, arquivo)
                zip_file.write(file_path, arquivo)

        zip_buffer.seek(0)

        # Gerar nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"conciliacao_{timestamp}.zip"

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Stats-Confirmados": str(resultado['stats']['confirmados']),
                "X-Stats-Previsao": str(resultado['stats']['previsao']),
                "X-Stats-Pagamentos": str(resultado['stats']['pagamentos']),
                "X-Stats-Transferencias": str(resultado['stats']['transferencias']),
            }
        )

    finally:
        # Limpar diretório temporário
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.get("/health")
async def health_check():
    """Endpoint de health check detalhado"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "dependencies": {
            "pandas": pd.__version__,
            "numpy": np.__version__
        }
    }


# ==============================================================================
# EXECUÇÃO
# ==============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=1909)
